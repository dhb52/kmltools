#!/usr/bin/env python
# coding: utf-8


import zipfile
from io import StringIO

import openpyxl
from fastkml import geometry, kml, styles
from geographiclib.geodesic import Geodesic
from shapely.geometry import LineString as ShapelyLineString
from shapely.geometry import Polygon as ShapelyPolygon
from shapely.geometry.polygon import Polygon as ShapelyPolygonType


class GeoPolygon:
    def __init__(self, name, path, coords):
        self.name = name
        self.path = path
        self.coords = coords


class GeoPoint:
    def __init__(self, name, path, coord):
        self.name = name
        self.path = path
        self.coord = coord


class GeoLine(object):
    def __init__(self, name, path, coords):
        self.name = name
        self.path = path
        self.coords = coords


def point_in_poly(x, y, poly):
    n = len(poly)
    inside = False
    p1x, p1y = poly[0]
    for i in range(n + 1):
        p2x, p2y = poly[i % n]
        if y > min(p1y, p2y):
            if y <= max(p1y, p2y):
                if x <= max(p1x, p2x):
                    if p1y != p2y:
                        xints = (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x
                    if p1x == p2x or x <= xints:
                        inside = not inside
        p1x, p1y = p2x, p2y
    return inside


def _find_features(result, element, feature_type, path=""):
    new_path = path
    if not getattr(element, "features", None):
        return
    for feature in element.features():
        if isinstance(feature, kml.Folder):
            if not feature.name:
                feature.name = ""
            new_path = "/".join((path, feature.name))
        elif isinstance(feature, kml.Placemark):
            if feature_type == geometry.Polygon and isinstance(
                feature.geometry, feature_type
            ):
                coords = [(p[0], p[1]) for p in feature.geometry.exterior.coords]
                poly = GeoPolygon(feature.name, path, coords)
                result.append(poly)
            elif feature_type == geometry.Point and isinstance(
                feature.geometry, feature_type
            ):
                point = GeoPoint(feature.name, path, feature.geometry)
                result.append(point)
            elif feature_type == geometry.LineString and isinstance(
                feature.geometry, feature_type
            ):
                line = GeoLine(feature.name, path, feature.geometry.coords)
                result.append(line)
        _find_features(result, feature, feature_type, new_path)


def _polygon_area(coordinates):
    geod = Geodesic.WGS84
    p = geod.Polygon()
    for lon, lat in coordinates:
        p.AddPoint(lat, lon)
    _, _, area = p.Compute()
    return abs(area)


def _read_kml_from_file(file_path):
    k = kml.KML()
    xml = None
    if file_path.lower().endswith(".kmz"):
        with zipfile.ZipFile(file_path, "r") as zf:
            xml = zf.read("doc.kml")
    else:
        with open(file_path, "rb") as kmlFile:
            xml = kmlFile.read()

    k.from_string(xml)
    return k


def _line_length(coords):
    geod = Geodesic.WGS84
    cnt = len(coords)
    length = 0.0
    for i in range(cnt - 1):
        p1_lon, p1_lat, _ = coords[i]
        p2_lon, p2_lat, _ = coords[i + 1]
        line = geod.Inverse(p1_lat, p1_lon, p2_lat, p2_lon)
        length += line["s12"]
    return length


def load_points(points_file):
    points = []
    k = _read_kml_from_file(points_file)
    _find_features(points, k, geometry.Point)
    return points


def load_polygons(polygons_file):
    polygons = []
    k = _read_kml_from_file(polygons_file)
    _find_features(polygons, k, geometry.Polygon)
    return polygons


def load_lines(line_file):
    lines = []
    k = _read_kml_from_file(line_file)
    _find_features(lines, k, geometry.LineString)
    return lines


def load_data(points_file, polygons_file):
    points = load_points(points_file)
    polygons = load_polygons(polygons_file)
    return points, polygons


def points_inside_info(points, polygons):
    result = StringIO()
    result.write("点名称,经度,纬度,点所在目录,多边形名称,多边形所在目录\n")
    for point in points:
        x, y = point.coord.x, point.coord.y
        for poly in polygons:
            if point_in_poly(x, y, poly.coords):
                result.write(
                    "%s,%f,%f,%s,%s,%s\n"
                    % (point.name, x, y, point.path, poly.name, poly.path)
                )
    return result.getvalue()


def calc_poly_areas(polygons):
    result = StringIO()
    result.write("名称,所在目录,面积(平方)\n")
    for a in polygons:
        area = _polygon_area(a.coords)
        result.write("%s,%s,%f\n" % (a.name, a.path, area))
    return result.getvalue()


def calc_line_length(lines):
    result = StringIO()
    result.write("名称,所在目录,长度(米)\n")
    for a in lines:
        length = _line_length(a.coords)
        result.write("%s,%s,%f\n" % (a.name, a.path, length))
    return result.getvalue()


def calc_polygon_intersection_area(polygons):
    result = StringIO()
    result.write("多边形1,多边形1所在目录,多边形2,多边形2所在目录,重叠面积(平方)\n")
    for i, p1 in enumerate(polygons):
        shapely_p1 = ShapelyPolygon(p1.coords)
        for p2 in polygons[i + 1 :]:
            shapely_p2 = ShapelyPolygon(p2.coords)
            inter = None
            try:
                inter = shapely_p1.intersection(shapely_p2)
            except:
                pass
            if isinstance(inter, ShapelyPolygonType):
                area = _polygon_area(inter.exterior.coords)
                if area > 0:
                    result.write(
                        "%s,%s,%s,%s,%f\n" % (p1.name, p1.path, p2.name, p2.path, area)
                    )

    return result.getvalue()


def points_to_csv(points):
    result = StringIO()
    result.write("名称,经度,纬度,点所在目录\n")
    for point in points:
        result.write(
            "%s,%f,%f,%s\n" % (point.name, point.coord.x, point.coord.y, point.path)
        )
    return result.getvalue()


def link_points_kml(excel_path, kml_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    kml_obj = kml.KML()
    lstyle = styles.LineStyle(color="ff00c500", width=2.0)
    style = styles.Style(styles=[lstyle])
    folder = kml.Folder(name="批量连线", styles=[style])

    kml_obj.append(folder)
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        p1_name, p1_lon, p1_lat, p2_name, p2_lon, p2_lat = row
        p1_lon, p1_lat, p2_lon, p2_lat = (
            float(p1_lon),
            float(p1_lat),
            float(p2_lon),
            float(p2_lat),
        )
        if p1_lon != p2_lon and p1_lat != p2_lat:
            line = ShapelyLineString([(p1_lon, p1_lat, 0), (p2_lon, p2_lat, 0)])
            p = kml.Placemark(name=f"{p1_name}<->{p2_name}", styles=[style])
            p.geometry = line
            folder.append(p)
    kmlstr = kml_obj.to_string(prettyprint=True)
    with open(kml_path, "w") as out:
        out.write(kmlstr)
    wb.close()


def save_excel_template(excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = (
        ("点1名称", "经度", "纬度", "点2名称", "经度", "纬度"),
        ("潮州市", 116.622604, 23.65695, "枫溪镇", 116.606779, 23.64921),
        ("潮州市", 116.622604, 23.65695, "湘桥区", 116.628632, 23.674536),
        ("潮州市", 116.622604, 23.65695, "古巷镇", 116.574761, 23.661714),
        ("潮州市", 116.622604, 23.65695, "东丽湖新村", 116.665511, 23.662787),
        ("潮州市", 116.622604, 23.65695, "下水头村", 116.630906, 23.623607),
        ("潮州市", 116.622604, 23.65695, "潮州市", 116.622604, 23.65695),
    )
    for row in rows:
        ws.append(row)
    wb.save(excel_path)
    wb.close()


if __name__ == "__main__":
    # points, polygons = load_data("points.kml", "areas.kml")
    # print(points_inside_info(points, polygons))
    # print(calc_poly_areas(polygons))
    save_excel_template("/Users/barry/Desktop/book.xlsx")
    link_points_kml("/Users/barry/Desktop/book.xlsx", "/Users/barry/Desktop/out.kml")
