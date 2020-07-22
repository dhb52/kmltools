from io import StringIO

import openpyxl
from fastkml import kml, styles
from shapely.geometry import LineString as ShapelyLineString


def calc_points_inside_polygons(points, polygons) -> str:
    result = StringIO()
    result.write("点名称,经度,纬度,点所在目录,多边形名称,多边形所在目录\n")
    for point in points:
        for poly in polygons:
            if poly.contains(point):
                result.write(
                    f"{point.name},{point.x},{point.y},{point.path},{poly.name},{poly.path}\n"
                )
    return result.getvalue()


def calc_polygons_area(polygons) -> str:
    result = StringIO()
    result.write("名称,所在目录,面积(平方)\n")
    for p in polygons:
        result.write(f"{p.name},{p.path},{p.area}\n")
    return result.getvalue()


def calc_lines_length(lines) -> str:
    result = StringIO()
    result.write("名称,所在目录,长度(米)\n")
    for line in lines:
        result.write(f"{line.name},{line.path},{line.length}\n")
    return result.getvalue()


def calc_polygons_intersection_area(polygons) -> str:
    result = StringIO()
    result.write("多边形1,多边形1所在目录,多边形2,多边形2所在目录,重叠面积(平方)\n")
    for i, p1 in enumerate(polygons):
        for p2 in polygons[i + 1 :]:
            area = p1.intersection_area(p2)
            if area > 0:
                result.write(f"{p1.name},{p1.path},{p2.name},{p2.path},{area}\n")

    return result.getvalue()


def points_to_csv(points) -> str:
    result = StringIO()
    result.write("名称,经度,纬度,点所在目录\n")
    for point in points:
        result.write(f"{point.name},{point.x},{point.y},{point.path}\n")
    return result.getvalue()


def link_points_kml(excel_path, kml_path) -> None:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    kml_obj = kml.KML()
    lstyle = styles.LineStyle(color="ff00c500", width=2.0)
    style = styles.Style(styles=[lstyle])
    folder = kml.Folder(name="批量连线", styles=[style])

    kml_obj.append(folder)
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        p1_name, p1_x, p1_y, p2_name, p2_x, p2_y = row
        p1_x, p1_y, p2_x, p2_y = (
            float(p1_x),
            float(p1_y),
            float(p2_x),
            float(p2_y),
        )
        if p1_x != p2_x and p1_y != p2_y:
            line = ShapelyLineString([(p1_x, p1_y, 0), (p2_x, p2_y, 0)])
            p = kml.Placemark(name=f"{p1_name}<->{p2_name}", styles=[style])
            p.geometry = line
            folder.append(p)
    kmlstr = kml_obj.to_string(prettyprint=True)
    with open(kml_path, "w", encoding="UTF8") as out:
        out.write(kmlstr)
    wb.close()


def save_excel_template(excel_path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = (
        ("点1名称", "经度", "纬度", "点2名称", "经度", "纬度"),
        ("潮州市", 116.622604, 23.65695, "枫溪镇", 116.606779, 23.64921),
        ("潮州市", 116.622604, 23.65695, "湘桥区", 116.628632, 23.674536),
        ("潮州市", 116.622604, 23.65695, "古巷镇", 116.574761, 23.661714),
        ("潮州市", 116.622604, 23.65695, "东丽湖新村", 116.665511, 23.662787),
        ("潮州市", 116.622604, 23.65695, "下水头村", 116.630906, 23.623607),
        ("潮州市", 116.622604, 23.65695, "潮州市", 116.622604, 23.65695),
    )
    for row in rows:
        ws.append(row)
    wb.save(excel_path)
    wb.close()


def link_radio_cran(radio_points, cran_points, grids, out_kml) -> str:
    grid_links = []
    for grid in grids:
        grid_cran = []
        grid_radio = []
        for c in cran_points:
            if grid.contains(c):
                grid_cran.append(c)
        for r in radio_points:
            if grid.contains(r):
                grid_radio.append(r)
        grid_links.append((grid, grid_cran, grid_radio))

    kml_obj = kml.KML()
    lstyle = styles.LineStyle(color="ff00c500", width=2.0)
    style = styles.Style(styles=[lstyle])
    root_folder = kml.Folder(name="CRAN连线", styles=[style])

    kml_obj.append(root_folder)
    result = StringIO()
    result.write(
        "归属网格,网格所在目录,CRAN机房,CRAN机房经度,CRAN机房纬度,CRAN机房所在目录,基站,基站经度,基站纬度,基站所在目录\n"
    )
    for link in grid_links:
        grid, grid_cran, grid_radio = link
        grid_folder = kml.Folder(name=grid.name)
        root_folder.append(grid_folder)
        for p1 in grid_cran:
            for p2 in grid_radio:
                if p1.x != p2.x and p1.y != p2.y:
                    result.write(
                        f"{grid.name},{grid.path},{p1.name},{p1.x},{p1.y},{p1.path},{p2.name},{p2.x},{p2.y},{p2.path}\n"
                    )
                    if len(grid_cran) == 1:
                        line = ShapelyLineString([(p1.x, p1.y, 0), (p2.x, p2.y, 0)])
                        p = kml.Placemark(name=f"{p1.name}<->{p2.name}", styles=[style])
                        p.geometry = line
                        grid_folder.append(p)

    kmlstr = kml_obj.to_string(prettyprint=True)
    with open(out_kml, mode="w", encoding="UTF8") as out:
        out.write(kmlstr)

    return result.getvalue()
